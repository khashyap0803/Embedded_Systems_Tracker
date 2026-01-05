#!/usr/bin/env python3
"""
Comprehensive merge of ODS (source of truth for quantities) with enriched Excel (metadata).

Rules:
1. Read ALL rows from ODS - this is the master list
2. For each ODS row, look up enriched data by COMPONENT name
3. If duplicate COMPONENT names in ODS, merge by summing quantities
4. Keep enriched metadata: common name, description, MPN, specs, category, datasheet
5. If no enriched data found, use ODS description as fallback
"""

from odf.opendocument import load
from odf.table import Table, TableRow, TableCell
from odf.text import P
from openpyxl import load_workbook
import json
from collections import defaultdict

def read_ods():
    """Read ODS file and return list of items (with duplicates)."""
    doc = load('COMPONENTS_DATA.ods')
    sheets = doc.spreadsheet.getElementsByType(Table)
    sheet = sheets[0]
    
    items = []
    for row in sheet.getElementsByType(TableRow)[1:]:
        cells = row.getElementsByType(TableCell)
        row_data = []
        for cell in cells:
            text_content = ''
            for p in cell.getElementsByType(P):
                text_content += str(p)
            repeat = cell.getAttribute('numbercolumnsrepeated')
            if repeat:
                for _ in range(min(int(repeat), 10)):
                    row_data.append(text_content.strip())
            else:
                row_data.append(text_content.strip())
        
        if len(row_data) >= 4 and row_data[0]:
            try:
                qty = int(row_data[3]) if row_data[3] else 1
            except:
                qty = 1
            
            items.append({
                'sno': row_data[0],
                'ods_description': row_data[1].strip() if row_data[1] else '',
                'component': row_data[2].strip().upper() if row_data[2] else '',
                'qty': qty
            })
    
    return items

def read_enriched():
    """Read enriched Excel and return dict keyed by original component name."""
    wb = load_workbook('COMPONENTS_DATA_ENRICHED.xlsx')
    sheet = wb.active
    
    data = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        
        # Column B is "Component Part Number" - this matches ODS COMPONENT column
        original = str(row[1]).strip().upper() if row[1] else ''
        
        data[original] = {
            'mpn': str(row[2]) if row[2] and row[2] != 'N/A' else None,
            'common_name': str(row[3]) if row[3] else None,
            'description': str(row[4]) if row[4] else None,
            'enriched_qty': row[5] if row[5] else 1,
            'package': str(row[6]) if row[6] else None,
            'specs': str(row[7]) if row[7] else None,
            'category': str(row[8]) if row[8] else 'Electronic Component',
            'datasheet': str(row[9]) if row[9] else None
        }
    
    return data

def merge_data():
    """Merge ODS quantities with enriched metadata."""
    ods_items = read_ods()
    enriched = read_enriched()
    
    print(f"ODS items: {len(ods_items)}")
    print(f"Enriched items: {len(enriched)}")
    
    # Step 1: Group ODS items by COMPONENT and sum quantities
    grouped = defaultdict(lambda: {
        'total_qty': 0,
        'ods_descriptions': [],
        'sno_list': []
    })
    
    for item in ods_items:
        comp = item['component']
        grouped[comp]['total_qty'] += item['qty']
        if item['ods_description'] and item['ods_description'] not in grouped[comp]['ods_descriptions']:
            grouped[comp]['ods_descriptions'].append(item['ods_description'])
        grouped[comp]['sno_list'].append(item['sno'])
    
    print(f"Unique components after grouping: {len(grouped)}")
    
    # Step 2: Create final components with enriched data
    final = []
    matched = 0
    unmatched = 0
    
    for i, (component, ods_data) in enumerate(sorted(grouped.items()), start=1):
        enriched_data = enriched.get(component)
        
        if enriched_data:
            matched += 1
            # Use enriched common name, fallback to component
            name = enriched_data['common_name'] or component.title()
            # Use enriched description, fallback to ODS description
            desc = enriched_data['description']
            if not desc and ods_data['ods_descriptions']:
                desc = '; '.join(ods_data['ods_descriptions'])
            
            comp = {
                'id': f'comp_{i:03d}',
                'original_name': component,
                'name': name,
                'mpn': enriched_data['mpn'],
                'description': desc,
                'quantity': ods_data['total_qty'],  # Use ODS quantity (merged)
                'package': enriched_data['package'],
                'specifications': enriched_data['specs'],
                'category': enriched_data['category'],
                'datasheet_url': enriched_data['datasheet'],
                'status': 'available'
            }
        else:
            unmatched += 1
            # Use ODS data only
            desc = '; '.join(ods_data['ods_descriptions']) if ods_data['ods_descriptions'] else component.title()
            comp = {
                'id': f'comp_{i:03d}',
                'original_name': component,
                'name': component.title(),
                'mpn': None,
                'description': desc,
                'quantity': ods_data['total_qty'],
                'package': None,
                'specifications': None,
                'category': 'Electronic Component',
                'datasheet_url': None,
                'status': 'available'
            }
            print(f"  UNMATCHED: {component} (qty={ods_data['total_qty']})")
        
        final.append(comp)
    
    print(f"\nMatched with enriched: {matched}")
    print(f"Unmatched (ODS only): {unmatched}")
    
    return final

def add_cart_items(components):
    """Add the 8 ordered cart items."""
    cart = [
        {
            'id': 'cart_001',
            'original_name': 'SIPEED SLOGIC COMBO 8',
            'name': 'Sipeed SLogic Combo 8 Logic Analyzer',
            'mpn': 'SLogic Combo 8 (Sipeed)',
            'description': 'The Daily Driver. 8-channel logic analyzer that decodes UART, SPI, I2C, PWM. Essential for firmware debugging.',
            'quantity': 1,
            'package': 'USB Device',
            'specifications': '8 channels, 24MHz, USB interface',
            'category': 'Tool / Measurement',
            'datasheet_url': None,
            'status': 'ordered',
            'price_inr': 1559
        },
        {
            'id': 'cart_002',
            'original_name': 'WEACT USB-CAN ANALYZER',
            'name': 'WeAct USB-CAN Analyzer (CAN-FD)',
            'mpn': 'USB-CAN Analyzer (WeAct)',
            'description': 'The Automotive Must-Have. Decodes CAN bus traffic for automotive embedded projects.',
            'quantity': 1,
            'package': 'USB Device',
            'specifications': 'CAN 2.0 + CAN-FD, USB interface',
            'category': 'Tool / Measurement',
            'datasheet_url': None,
            'status': 'ordered',
            'price_inr': 1075
        },
        {
            'id': 'cart_003',
            'original_name': 'SN65HVD230 CAN BOARD',
            'name': 'SN65HVD230 CAN Transceiver Board',
            'mpn': 'SN65HVD230 (Texas Instruments)',
            'description': 'Connects STM32 to CAN Bus. 3.3V CAN transceiver module.',
            'quantity': 2,
            'package': 'Module (SOP-8 IC)',
            'specifications': '3.3V, CAN 2.0, 1Mbps',
            'category': 'Bus Transceiver',
            'datasheet_url': 'https://www.ti.com/lit/ds/symlink/sn65hvd230.pdf',
            'status': 'ordered',
            'price_inr': 418
        },
        {
            'id': 'cart_004',
            'original_name': 'MCP2003-E/P LIN TRANSCEIVER',
            'name': 'MCP2003-E/P LIN Transceiver',
            'mpn': 'MCP2003-E/P (Microchip)',
            'description': 'Connects STM32 to LIN Bus for automotive networks.',
            'quantity': 2,
            'package': 'DIP-8',
            'specifications': 'LIN 2.x, 20kbps, sleep mode',
            'category': 'Bus Transceiver',
            'datasheet_url': 'https://ww1.microchip.com/downloads/en/DeviceDoc/20002230G.pdf',
            'status': 'ordered',
            'price_inr': 152
        },
        {
            'id': 'cart_005',
            'original_name': 'ENC28J60 ETHERNET MODULE',
            'name': 'ENC28J60 Ethernet Module',
            'mpn': 'ENC28J60 (Microchip)',
            'description': 'For learning TCP/IP stacks. SPI-based Ethernet controller.',
            'quantity': 1,
            'package': 'Module',
            'specifications': '10Base-T, SPI interface, 8KB buffer',
            'category': 'Ethernet Module',
            'datasheet_url': 'https://ww1.microchip.com/downloads/en/DeviceDoc/39662e.pdf',
            'status': 'ordered',
            'price_inr': 299
        },
        {
            'id': 'cart_006',
            'original_name': 'TXS0108E LEVEL CONVERTER',
            'name': 'TXS0108E Logic Level Converter',
            'mpn': 'TXS0108E (Texas Instruments)',
            'description': 'Safety converter for 5V <-> 3.3V bidirectional level shifting.',
            'quantity': 1,
            'package': 'Module (TSSOP-20)',
            'specifications': '8-channel, bidirectional, auto-direction',
            'category': 'Logic Level Converter',
            'datasheet_url': 'https://www.ti.com/lit/ds/symlink/txs0108e.pdf',
            'status': 'ordered',
            'price_inr': 48
        },
        {
            'id': 'cart_007',
            'original_name': 'INA219 CURRENT SENSOR',
            'name': 'INA219 Current Sensor Module',
            'mpn': 'INA219 (Texas Instruments)',
            'description': 'For measuring power consumption. I2C current/voltage sensor.',
            'quantity': 1,
            'package': 'Module (SOT23-8)',
            'specifications': 'I2C, 0-26V, ±3.2A, 12-bit',
            'category': 'Sensor Module',
            'datasheet_url': 'https://www.ti.com/lit/ds/symlink/ina219.pdf',
            'status': 'ordered',
            'price_inr': 200
        },
        {
            'id': 'cart_008',
            'original_name': 'USB MICROSCOPE 1000X',
            'name': 'USB Digital Microscope 1000X',
            'mpn': None,
            'description': 'For inspecting solder joints, SMD components, and PCB traces.',
            'quantity': 1,
            'package': 'USB Device',
            'specifications': '1000X magnification, LED lights, USB camera',
            'category': 'Tool / Measurement',
            'datasheet_url': None,
            'status': 'ordered',
            'price_inr': None
        }
    ]
    
    components.extend(cart)
    return components

def main():
    # Merge ODS + enriched data
    components = merge_data()
    
    # Add cart items
    components = add_cart_items(components)
    
    # Build inventory
    total_qty = sum(c['quantity'] for c in components)
    
    inventory = {
        'metadata': {
            'version': '3.0',
            'last_updated': '2026-01-05',
            'owner': 'Nani',
            'notes': 'Merged from COMPONENTS_DATA.ods (quantities) + COMPONENTS_DATA_ENRICHED.xlsx (metadata) + 8 ordered cart items'
        },
        'components': components,
        'summary': {
            'total_unique_components': len(components),
            'total_quantity': total_qty,
            'available': sum(1 for c in components if c['status'] == 'available'),
            'ordered': sum(1 for c in components if c['status'] == 'ordered')
        }
    }
    
    # Save
    with open('embedded_tracker/data/hardware_inventory.json', 'w') as f:
        json.dump(inventory, f, indent=4)
    
    print(f"\n=== FINAL INVENTORY ===")
    print(f"Total unique components: {len(components)}")
    print(f"Total quantity: {total_qty}")
    print(f"Available: {inventory['summary']['available']}")
    print(f"Ordered: {inventory['summary']['ordered']}")
    
    # Show sample with all fields
    print("\n=== SAMPLE ENTRIES ===")
    for c in components[:3]:
        print(f"\n{c['name']}")
        print(f"  Original: {c['original_name']}")
        print(f"  MPN: {c['mpn']}")
        print(f"  Qty: {c['quantity']}")
        print(f"  Category: {c['category']}")
        print(f"  Package: {c['package']}")
        print(f"  Specs: {c['specifications']}")
        print(f"  Desc: {c['description'][:80] if c['description'] else None}...")
        print(f"  Datasheet: {c['datasheet_url']}")

if __name__ == '__main__':
    main()
